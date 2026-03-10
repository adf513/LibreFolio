#!/usr/bin/env python3
"""
i18n Audit Tool for LibreFolio

This script analyzes translation files and generates a comparison table
showing all translation keys across all languages, highlighting missing translations.

Usage:
    python scripts/i18n-audit.py [--format md|xlsx|both] [--output path]

Examples:
    python scripts/i18n-audit.py                    # Markdown to stdout
    python scripts/i18n-audit.py --format xlsx     # Excel file in current dir
    python scripts/i18n-audit.py --format both     # Both formats in current dir
    python scripts/i18n-audit.py -o ./reports/     # Output to specific directory
"""

import json
import sys
from pathlib import Path
from typing import Any

# Try to import required libraries
try:
    import pandas as pd
except ImportError:
    print("❌ pandas is required for this script.")
    print("   Install with: pipenv install pandas")
    sys.exit(1)

try:
    from tabulate import tabulate
except ImportError:
    print("❌ tabulate is required for this script.")
    print("   Install with: pipenv install tabulate --dev")
    sys.exit(1)

# Configuration
I18N_DIR = Path(__file__).parent.parent / "src" / "lib" / "i18n"
LANGUAGES = ["en", "it", "fr", "es"]  # Order matters for display


def flatten_dict(d: dict, parent_key: str = "", sep: str = ".") -> dict[str, Any]:
    """
    Flatten a nested dictionary into dot-notation keys.

    Example:
        {"common": {"loading": "Loading..."}} -> {"common.loading": "Loading..."}
    """
    items: list[tuple[str, Any]] = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep).items())
        else:
            items.append((new_key, v))
    return dict(items)


def find_used_keys_in_sources() -> tuple[set[str], set[str]]:
    """
    Scan all .svelte and .ts files in src/ to find translation keys being used.

    Returns:
        Tuple of (exact_keys, prefix_patterns)
        - exact_keys: Keys that are definitely used (full path found)
        - prefix_patterns: Prefixes that might be used dynamically (e.g., 'fileStatus')

    Looks for patterns like:
    - $t('key.path') / t('key.path') / $_('key.path')
    - labelKey: 'nav.dashboard' (keys passed as props)
    - `prefix.${var}` (template literal interpolation)
    """
    import re

    src_dir = I18N_DIR.parent.parent  # src/
    exact_keys: set[str] = set()
    prefix_patterns: set[str] = set()

    # Patterns to match DIRECT translation function calls
    patterns_exact = [
        # $t('key') or $_('key') or t('key') with single/double quotes
        re.compile(r"(?:\$t|\$_|(?<![a-zA-Z])t|(?<![a-zA-Z])_)\s*\(\s*['\"]([a-zA-Z0-9_.]+)['\"]"),
        ]

    # Patterns for INDIRECT keys (passed as props/variables)
    # These are keys stored in objects/arrays and later passed to $t/$_
    patterns_indirect = [
        # labelKey: 'nav.dashboard', titleKey: 'settings.title', etc.
        re.compile(r"(?:label|title|text|message|hint|description|placeholder)Key['\"]?\s*[:=]\s*['\"]([a-zA-Z0-9_.]+)['\"]"),
        # key: 'common.save' in objects
        re.compile(r"['\"]?key['\"]?\s*:\s*['\"]([a-zA-Z0-9_.]+)['\"]"),
        # i18n key strings in arrays: ['common.yes', 'common.no']
        re.compile(r"\[\s*['\"]([a-zA-Z0-9_.]+)['\"](?:\s*,\s*['\"]([a-zA-Z0-9_.]+)['\"])*\s*\]"),
        ]

    # Patterns for DYNAMIC keys (template literals / concatenation)
    patterns_dynamic = [
        # `prefix.${...}` - extract the prefix
        re.compile(r"`([a-zA-Z0-9_.]+)\.\$\{"),
        # 'prefix.' + var or "prefix." + var
        re.compile(r"['\"]([a-zA-Z0-9_.]+)\.['\"]?\s*\+"),
        ]

    # Scan all .svelte, .ts, and .js files
    for ext in ["*.svelte", "*.ts", "*.js"]:
        for file_path in src_dir.rglob(ext):
            # Skip node_modules and build directories
            if "node_modules" in str(file_path) or "/build/" in str(file_path):
                continue
            try:
                content = file_path.read_text(encoding="utf-8")

                # Find exact keys (direct calls)
                for pattern in patterns_exact:
                    matches = pattern.findall(content)
                    exact_keys.update(matches)

                # Find indirect keys (passed as props)
                for pattern in patterns_indirect:
                    matches = pattern.findall(content)
                    # Flatten tuples from regex groups
                    for match in matches:
                        if isinstance(match, tuple):
                            exact_keys.update(m for m in match if m)
                        else:
                            exact_keys.add(match)

                # Find dynamic prefixes
                for pattern in patterns_dynamic:
                    matches = pattern.findall(content)
                    prefix_patterns.update(matches)

            except Exception:
                pass  # Skip files that can't be read

    return exact_keys, prefix_patterns


def is_key_potentially_used(key: str, exact_keys: set[str], prefix_patterns: set[str]) -> bool:
    """
    Check if a key is potentially used, either exactly or via dynamic construction.
    """
    # Exact match
    if key in exact_keys:
        return True

    # Check if any prefix pattern matches the start of this key
    for prefix in prefix_patterns:
        if key.startswith(prefix + "."):
            return True

    return False


def load_translations() -> dict[str, dict[str, str]]:
    """
    Load all translation files and return flattened dictionaries.

    Returns:
        Dict mapping language code to flattened translations
    """
    translations: dict[str, dict[str, str]] = {}

    for lang in LANGUAGES:
        file_path = I18N_DIR / f"{lang}.json"
        if file_path.exists():
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                translations[lang] = flatten_dict(data)
                print(f"✓ Loaded {lang}.json ({len(translations[lang])} keys)")
            except json.JSONDecodeError as e:
                print(f"❌ Error parsing {lang}.json: {e}")
                translations[lang] = {}
        else:
            print(f"⚠ File not found: {lang}.json")
            translations[lang] = {}

    return translations


def get_all_keys(translations: dict[str, dict[str, str]]) -> list[str]:
    """
    Get all unique keys across all languages, sorted.
    """
    all_keys: set[str] = set()
    for lang_data in translations.values():
        all_keys.update(lang_data.keys())
    return sorted(all_keys)


def extract_section(key: str) -> str:
    """
    Extract the top-level section from a dotted key.

    Example: "common.loading" -> "common"
    """
    return key.split(".")[0] if "." in key else key


def build_dataframe(translations: dict[str, dict[str, str]]) -> pd.DataFrame:
    """
    Build a DataFrame with all translations.

    Columns: Section, Key, en, it, fr, es, Status
    """
    all_keys = get_all_keys(translations)

    rows: list[dict[str, Any]] = []
    for key in all_keys:
        row: dict[str, Any] = {
            "Section": extract_section(key),
            "Key": key,
            }

        # Add each language value
        missing_count = 0
        for lang in LANGUAGES:
            value = translations.get(lang, {}).get(key)
            row[lang.upper()] = value if value else ""
            if not value:
                missing_count += 1

        # Status column
        if missing_count == 0:
            row["Status"] = "✅"
        elif missing_count == len(LANGUAGES):
            row["Status"] = "❌ MISSING ALL"
        else:
            missing_langs = [lang.upper() for lang in LANGUAGES
                             if not translations.get(lang, {}).get(key)]
            row["Status"] = f"⚠️ Missing: {', '.join(missing_langs)}"

        rows.append(row)

    return pd.DataFrame(rows)


def generate_summary(df: pd.DataFrame) -> str:
    """
    Generate a summary of translation coverage.
    """
    lines = [
        "## 📊 Translation Coverage Summary\n",
        f"**Total keys**: {len(df)}\n",
        ]

    # Per-language stats
    for lang in LANGUAGES:
        col = lang.upper()
        filled = df[col].apply(lambda x: bool(x)).sum()
        pct = (filled / len(df)) * 100 if len(df) > 0 else 0
        lines.append(f"- **{col}**: {filled}/{len(df)} ({pct:.1f}%)")

    # Missing translations
    incomplete = df[df["Status"] != "✅"]
    if len(incomplete) > 0:
        lines.append(f"\n**⚠️ Incomplete translations**: {len(incomplete)}\n")
    else:
        lines.append("\n**✅ All translations complete!**\n")

    return "\n".join(lines)


def generate_missing_report(df: pd.DataFrame) -> str:
    """
    Generate a report of missing translations using tabulate for nice formatting.
    """
    incomplete = df[df["Status"] != "✅"].copy()

    if len(incomplete) == 0:
        return "\n## ✅ No Missing Translations\n\nAll keys are translated in all languages.\n"

    lines = [
        "\n## ⚠️ Missing Translations\n",
        "The following keys need attention:\n",
        ]

    # Group by section
    for section in incomplete["Section"].unique():
        section_rows = incomplete[incomplete["Section"] == section]
        lines.append(f"\n### {section}\n")

        # Build table data for tabulate
        table_data = []
        for _, row in section_rows.iterrows():
            table_data.append([f"`{row['Key']}`", row['Status']])

        # Use tabulate for nice formatting
        table_str = tabulate(
            table_data,
            headers=["Key", "Status"],
            tablefmt="github"
            )
        lines.append(table_str)
        lines.append("")

    return "\n".join(lines)


def generate_unused_keys_report(
    all_keys: list[str],
    exact_keys: set[str],
    prefix_patterns: set[str]
    ) -> tuple[str, list[str]]:
    """
    Generate a report of translation keys that exist but are not used in the codebase.

    Returns:
        Tuple of (report string, list of unused keys)
    """
    unused_keys = [
        key for key in all_keys
        if not is_key_potentially_used(key, exact_keys, prefix_patterns)
        ]

    if len(unused_keys) == 0:
        return "\n## ✅ No Unused Translation Keys\n\nAll translation keys are used in the codebase.\n", []

    lines = [
        f"\n## ⚠️ Potentially Unused Translation Keys ({len(unused_keys)})\n",
        "The following keys exist in translation files but were not found in source code.\n",
        "\n**Note:** This analysis cannot detect:\n",
        "- Keys constructed dynamically (e.g., `$t(\\`prefix.${var}\\`)`)\n",
        "- Keys passed as variables\n",
        "- Keys used in computed expressions\n\n",
        ]

    if prefix_patterns:
        lines.append(f"**Dynamic prefixes detected:** `{', '.join(sorted(prefix_patterns))}`\n")
        lines.append("Keys under these prefixes are marked as potentially used.\n\n")

    # Group by section
    sections: dict[str, list[str]] = {}
    for key in unused_keys:
        section = extract_section(key)
        if section not in sections:
            sections[section] = []
        sections[section].append(key)

    for section in sorted(sections.keys()):
        lines.append(f"\n### {section}\n")
        for key in sorted(sections[section]):
            lines.append(f"- `{key}`")
        lines.append("")

    return "\n".join(lines), unused_keys


def export_markdown(
    df: pd.DataFrame,
    output_path: Path | None,
    include_full_table: bool = True,
    unused_keys_report: str = ""
    ) -> None:
    """
    Export the translation table to Markdown format using tabulate.

    Args:
        df: DataFrame with translations
        output_path: Output file path, or None for stdout
        include_full_table: If False, only show summary and missing translations
        unused_keys_report: Report of unused translation keys
    """
    summary = generate_summary(df)
    missing_report = generate_missing_report(df)

    # Prepare DataFrame for tabulate (truncate long values)
    display_df = df.copy()
    for col in display_df.columns:
        display_df[col] = display_df[col].apply(
            lambda x: (str(x)[:10] + "...") if len(str(x)) > 13 else str(x)
            )

    # Generate markdown table using tabulate
    table_md = tabulate(
        display_df,
        headers="keys",
        tablefmt="github",  # GitHub-flavored markdown
        showindex=False
        )

    # Build report parts
    parts = [
        "# LibreFolio i18n Audit Report\n",
        f"*Generated: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')}*\n",
        summary,
        missing_report,
        ]

    if unused_keys_report:
        parts.append(unused_keys_report)

    if include_full_table:
        parts.extend([
            "\n## 📋 Complete Translation Table\n",
            table_md,
            ])

    parts.append("")
    content = "\n".join(parts)

    # Write to file or stdout
    if output_path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"\n✅ Markdown report saved to: {output_path}")
    else:
        print(content)


def export_excel(df: pd.DataFrame, output_path: Path) -> None:
    """
    Export the translation table to Excel format with formatting.
    """
    try:
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("❌ openpyxl is required for Excel export.")
        print("   Install with: pip install openpyxl")
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Create Excel writer with formatting
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Main sheet with all translations
        df.to_excel(writer, sheet_name="All Translations", index=False)

        # Incomplete translations sheet
        incomplete = df[df["Status"] != "✅"]
        if len(incomplete) > 0:
            incomplete.to_excel(writer, sheet_name="Missing", index=False)

        # Summary sheet
        summary_data = {
            "Language": [lang.upper() for lang in LANGUAGES],
            "Filled": [df[lang.upper()].apply(lambda x: bool(x)).sum() for lang in LANGUAGES],
            "Missing": [df[lang.upper()].apply(lambda x: not bool(x)).sum() for lang in LANGUAGES],
            "Coverage %": [
                f"{(df[lang.upper()].apply(lambda x: bool(x)).sum() / len(df)) * 100:.1f}%"
                for lang in LANGUAGES
                ],
            }
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Apply formatting
        workbook = writer.book

        # Format header row
        header_fill = PatternFill(start_color="1A4D3E", end_color="1A4D3E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

    print(f"\n✅ Excel report saved to: {output_path}")


def run_audit(format_type: str = "none", output: str | None = None) -> int:
    """Run the i18n audit with specified format and output."""
    # Determine output directory (PWD by default)
    if output:
        output_base = Path(output)
        if output_base.suffix in [".md", ".xlsx"]:
            # User specified a file, use its parent as base
            output_dir = output_base.parent
        else:
            output_dir = output_base
    else:
        output_dir = Path.cwd()

    print("=" * 60)
    print("  LibreFolio i18n Audit Tool")
    print("=" * 60)
    print()

    # Load translations
    print("📂 Loading translation files...")
    translations = load_translations()
    print()

    # Build DataFrame
    print("🔍 Analyzing translations...")
    df = build_dataframe(translations)

    # Find unused keys
    print("🔎 Scanning source files for used keys...")
    all_keys = get_all_keys(translations)
    exact_keys, prefix_patterns = find_used_keys_in_sources()
    unused_report, unused_keys = generate_unused_keys_report(all_keys, exact_keys, prefix_patterns)
    print(f"   Found {len(exact_keys)} exact keys in source code")
    print(f"   Found {len(prefix_patterns)} dynamic prefixes: {', '.join(sorted(prefix_patterns)) or 'none'}")
    print(f"   Found {len(unused_keys)} potentially unused keys")
    print()

    # Generate output based on format
    if format_type == "none":
        # Only show summary and warnings, no full table
        export_markdown(df, None, include_full_table=False, unused_keys_report=unused_report)
    elif format_type in ["md", "both"]:
        if output and output.endswith(".md"):
            md_path = Path(output)
        elif format_type == "md" and output is None:
            # Print to stdout if no output specified for md-only
            export_markdown(df, None, include_full_table=True, unused_keys_report=unused_report)
            md_path = None
        else:
            md_path = output_dir / "i18n-audit.md"

        if md_path:
            export_markdown(df, md_path, include_full_table=True, unused_keys_report=unused_report)

    if format_type in ["xlsx", "both"]:
        if output and output.endswith(".xlsx"):
            xlsx_path = Path(output)
        else:
            xlsx_path = output_dir / "i18n-audit.xlsx"
        export_excel(df, xlsx_path)

    # Print summary
    print("\n" + "=" * 60)
    print("  Summary")
    print("=" * 60)
    print(f"  Total keys: {len(df)}")
    complete = len(df[df["Status"] == "✅"])
    incomplete = len(df) - complete
    print(f"  Complete:   {complete} ✅")
    print(f"  Incomplete: {incomplete} {'⚠️' if incomplete > 0 else ''}")
    print(f"  Unused:     {len(unused_keys)} {'⚠️' if len(unused_keys) > 0 else ''}")
    print()

    return 0


def add_arguments(parser) -> None:
    """Add arguments to a parser (reusable for both standalone and subparser)."""
    parser.add_argument(
        "--format", "-f",
        choices=["none", "md", "xlsx", "both"],
        default="none",
        help="Output format: none=only warnings (default), md=full markdown, xlsx=excel, both=md+xlsx"
        )
    parser.add_argument(
        "--output", "-o",
        type=str,
        default=None,
        help="Output directory or file path (default: current directory for file output)"
        )


def run_from_args(args) -> int:
    """Execute the command from parsed args."""
    return run_audit(
        format_type=getattr(args, 'format', 'none'),
        output=getattr(args, 'output', None)
        )


def main():
    """Main entry point."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Audit LibreFolio i18n translations",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python i18n-audit.py                    Show warnings only (no full table)
  python i18n-audit.py --format md        Show full Markdown report in terminal
  python i18n-audit.py --format xlsx      Generate Excel file in current dir
  python i18n-audit.py --format both      Generate both formats in current dir
  python i18n-audit.py -o ./reports/      Save to specific directory
        """
        )
    add_arguments(parser)
    args = parser.parse_args()
    return run_from_args(args)


# =============================================================================
# i18n CLI Management Functions (add, remove, update, search)
# =============================================================================

def get_nested_key(data: dict, key: str) -> str | None:
    """Get value from nested dict using dot notation key."""
    parts = key.split(".")
    current = data
    for part in parts:
        if isinstance(current, dict) and part in current:
            current = current[part]
        else:
            return None
    return current if isinstance(current, str) else None


def set_nested_key(data: dict, key: str, value: str) -> bool:
    """Set value in nested dict using dot notation key. Creates parents if needed."""
    parts = key.split(".")
    current = data
    for part in parts[:-1]:
        if part not in current:
            current[part] = {}
        elif not isinstance(current[part], dict):
            return False  # Cannot create nested key, parent is not a dict
        current = current[part]
    current[parts[-1]] = value
    return True


def delete_nested_key(data: dict, key: str) -> bool:
    """Delete key from nested dict using dot notation. Cleans up empty parents."""
    parts = key.split(".")

    # Navigate to parent
    current = data
    parents = [(data, None)]  # Stack of (dict, key_in_parent)

    for part in parts[:-1]:
        if isinstance(current, dict) and part in current:
            parents.append((current, part))
            current = current[part]
        else:
            return False

    # Delete the final key
    if parts[-1] in current:
        del current[parts[-1]]

        # Clean up empty parent dictionaries
        for parent_dict, parent_key in reversed(parents[1:]):
            if parent_key and isinstance(parent_dict[parent_key], dict) and len(parent_dict[parent_key]) == 0:
                del parent_dict[parent_key]

        return True
    return False


def load_lang_file(lang: str) -> dict:
    """Load translation file for a language."""
    filepath = I18N_DIR / f"{lang}.json"
    with open(filepath, "r", encoding="utf-8") as f:
        return json.load(f)


def save_lang_file(lang: str, data: dict) -> None:
    """Save translation file for a language with consistent formatting."""
    filepath = I18N_DIR / f"{lang}.json"
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")  # Trailing newline


def cmd_add(args) -> int:
    """Add a new translation key to all languages."""
    key = args.key
    translations = {
        "en": args.en,
        "it": args.it,
        "fr": args.fr,
        "es": args.es,
        }

    # Check if key already exists in any language
    existing_values = {}
    for lang in LANGUAGES:
        data = load_lang_file(lang)
        value = get_nested_key(data, key)
        if value is not None:
            existing_values[lang] = value

    if existing_values:
        print(f"\n❌ Key '{key}' already exists:\n")
        # Build table row
        row = [key]
        for lang in LANGUAGES:
            if lang in existing_values:
                row.append(f"✗ {existing_values[lang]}")
            else:
                row.append("○ (missing)")

        headers = ["Key"] + [lang.upper() for lang in LANGUAGES]
        print(tabulate([row], headers=headers, tablefmt="simple"))
        print(f"\n💡 Use 'i18n update' to modify existing translations")
        return 1

    # Add to all languages
    success_row = [key]
    for lang in LANGUAGES:
        data = load_lang_file(lang)
        if set_nested_key(data, key, translations[lang]):
            save_lang_file(lang, data)
            success_row.append(f"✓ {translations[lang]}")
        else:
            print(f"❌ Failed to add '{key}' to {lang}.json (parent path conflict)")
            return 1

    print(f"\n✅ Successfully added key '{key}':\n")
    headers = ["Key"] + [lang.upper() for lang in LANGUAGES]
    print(tabulate([success_row], headers=headers, tablefmt="simple"))
    return 0


def cmd_remove(args) -> int:
    """Remove a translation key from all languages."""
    key = args.key

    # Check current values
    current_values = {}
    for lang in LANGUAGES:
        data = load_lang_file(lang)
        value = get_nested_key(data, key)
        if value is not None:
            current_values[lang] = value

    if not current_values:
        print(f"\n⚠️  Key '{key}' was not found in any language")
        return 0

    if not args.force:
        # Show current values as table before confirming
        print(f"\n🔍 Current values for '{key}':\n")
        row = [key]
        for lang in LANGUAGES:
            if lang in current_values:
                row.append(current_values[lang])
            else:
                row.append("—")
        headers = ["Key"] + [lang.upper() for lang in LANGUAGES]
        print(tabulate([row], headers=headers, tablefmt="simple"))
        print()

        try:
            confirm = input(f"Remove '{key}' from all languages? [y/N]: ")
        except EOFError:
            print("\nAborted.")
            return 0
        if confirm.lower() != 'y':
            print("Aborted.")
            return 0

    # Remove from all languages
    result_row = [key]
    removed_count = 0
    for lang in LANGUAGES:
        data = load_lang_file(lang)
        if delete_nested_key(data, key):
            save_lang_file(lang, data)
            result_row.append("✓ removed")
            removed_count += 1
        else:
            result_row.append("— (not found)")

    print(f"\n✅ Removed key '{key}':\n")
    headers = ["Key"] + [lang.upper() for lang in LANGUAGES]
    print(tabulate([result_row], headers=headers, tablefmt="simple"))
    return 0


def cmd_update(args) -> int:
    """Update a specific translation in one or more languages."""
    key = args.key
    updates = {}

    if args.en: updates["en"] = args.en
    if args.it: updates["it"] = args.it
    if args.fr: updates["fr"] = args.fr
    if args.es: updates["es"] = args.es

    if not updates:
        print("❌ At least one language must be specified (--en, --it, --fr, --es)")
        return 1

    # Get current values for all languages
    current_values = {}
    for lang in LANGUAGES:
        data = load_lang_file(lang)
        value = get_nested_key(data, key)
        if value is not None:
            current_values[lang] = value

    if not current_values:
        print(f"❌ Key '{key}' does not exist. Use 'i18n add' to create it first.")
        return 1

    # Update specified languages and build result table
    result_row = [key]
    for lang in LANGUAGES:
        old_value = current_values.get(lang)
        if lang in updates:
            new_value = updates[lang]
            data = load_lang_file(lang)
            if set_nested_key(data, key, new_value):
                save_lang_file(lang, data)
                if old_value:
                    result_row.append(f"✓ {old_value} → {new_value}")
                else:
                    result_row.append(f"✓ (new) {new_value}")
            else:
                print(f"❌ Failed to update {lang}.json")
                return 1
        else:
            # Not updated, show current value
            if old_value:
                result_row.append(f"— {old_value}")
            else:
                result_row.append("— (missing)")

    print(f"\n✅ Updated key '{key}':\n")
    headers = ["Key"] + [lang.upper() for lang in LANGUAGES]
    print(tabulate([result_row], headers=headers, tablefmt="simple", maxcolwidths=[40, 30, 30, 30, 30]))
    return 0
    return 0


def cmd_search(args) -> int:
    """Search for keys/values in translations.

    Supports filtering by:
    - --keys: search only in key names
    - --values: search only in translation values
    - --lang: restrict value search to specific language(s)
    Default (no flags): search in keys AND all values.
    Result always shows ALL languages for matching keys.
    """
    query = args.query.lower()
    search_keys = getattr(args, 'keys', False)
    search_values = getattr(args, 'values', False)
    lang_filter = getattr(args, 'lang', None)  # e.g. "en,it"

    # Default: search both keys and values if neither flag specified
    if not search_keys and not search_values:
        search_keys = True
        search_values = True

    # Parse language filter
    search_langs = None
    if lang_filter:
        search_langs = [l.strip().lower() for l in lang_filter.split(",") if l.strip()]
        # Validate
        for sl in search_langs:
            if sl not in LANGUAGES:
                print(f"❌ Unknown language '{sl}'. Available: {', '.join(LANGUAGES)}")
                return 1

    # Load all language data
    all_data = {}
    all_flat = {}
    for lang in LANGUAGES:
        data = load_lang_file(lang)
        all_data[lang] = data
        all_flat[lang] = flatten_dict(data)

    # Collect all known keys (union of all languages)
    all_keys = set()
    for lang in LANGUAGES:
        all_keys.update(all_flat[lang].keys())

    # Find matching keys
    matching_keys = set()

    for key in all_keys:
        matched = False

        # Search in key name
        if search_keys and query in key.lower():
            matched = True

        # Search in values
        if search_values and not matched:
            langs_to_check = search_langs if search_langs else LANGUAGES
            for lang in langs_to_check:
                value = all_flat[lang].get(key, "")
                if isinstance(value, str) and query in value.lower():
                    matched = True
                    break

        if matched:
            matching_keys.add(key)

    if not matching_keys:
        print(f"🔍 No results found for '{args.query}'")
        return 0

    # Build table rows — ALWAYS show ALL languages
    rows = []
    for key in sorted(matching_keys):
        row = [key]
        for lang in LANGUAGES:
            value = all_flat[lang].get(key)
            if value is not None and isinstance(value, str):
                row.append(value)
            else:
                row.append("(missing)")
        rows.append(row)

    # Build mode description for output
    mode_parts = []
    if search_keys:
        mode_parts.append("keys")
    if search_values:
        if search_langs:
            mode_parts.append(f"values[{','.join(search_langs)}]")
        else:
            mode_parts.append("values[all]")
    mode_str = " + ".join(mode_parts)

    print(f"\n🔍 Found {len(matching_keys)} key(s) matching '{args.query}' (in {mode_str}):\n")
    headers = ["Key"] + [lang.upper() for lang in LANGUAGES]
    print(tabulate(rows, headers=headers, tablefmt="simple", maxcolwidths=[40, 30, 30, 30, 30]))

    return 0


def cmd_tree(args) -> int:
    """Show translation key structure as an ASCII tree.

    Displays the nested JSON structure of translation keys, useful
    for understanding the i18n namespace hierarchy at a glance.

    Usage:
        ./dev.py i18n tree                     # Full tree
        ./dev.py i18n tree chartSettings       # Only chartSettings subtree
        ./dev.py i18n tree -d 2                # Limit depth to 2 levels
        ./dev.py i18n tree --counts            # Show leaf count per branch
    """
    prefix = getattr(args, 'prefix', '') or ''
    max_depth = getattr(args, 'depth', 0) or 0
    show_counts = getattr(args, 'counts', False)

    # Load English as the reference structure
    data = load_lang_file("en")

    # Navigate to prefix if specified
    if prefix:
        parts = prefix.split(".")
        for part in parts:
            if isinstance(data, dict) and part in data:
                data = data[part]
            else:
                print(f"❌ Prefix '{prefix}' not found in en.json")
                return 1
        if not isinstance(data, dict):
            print(f"'{prefix}' is a leaf node: \"{data}\"")
            return 0

    def count_leaves(d: dict | str) -> int:
        """Count leaf (string) nodes recursively."""
        if isinstance(d, str):
            return 1
        return sum(count_leaves(v) for v in d.values())

    def print_tree(d: dict, indent: str = "", depth: int = 1) -> None:
        """Print dict as an ASCII tree with ├── and └── connectors."""
        keys = list(d.keys())
        for i, key in enumerate(keys):
            is_last = (i == len(keys) - 1)
            connector = "└── " if is_last else "├── "
            child = d[key]

            if isinstance(child, dict):
                count_str = f"  ({count_leaves(child)} keys)" if show_counts else ""
                print(f"{indent}{connector}📂 {key}{count_str}")
                if max_depth == 0 or depth < max_depth:
                    next_indent = indent + ("    " if is_last else "│   ")
                    print_tree(child, next_indent, depth + 1)
                elif max_depth > 0:
                    next_indent = indent + ("    " if is_last else "│   ")
                    print(f"{next_indent}└── ...")
            else:
                # Leaf node — show truncated value
                val_str = str(child)
                if len(val_str) > 50:
                    val_str = val_str[:47] + "..."
                print(f"{indent}{connector}🔑 {key}: \"{val_str}\"")

    # Header
    total_leaves = count_leaves(data)
    if prefix:
        print(f"\n🌳 i18n tree for '{prefix}' ({total_leaves} keys):\n")
    else:
        print(f"\n🌳 i18n tree ({total_leaves} keys):\n")

    print_tree(data)
    print()
    return 0


def register_subparser(subparsers) -> None:
    """Register as subparser for dev.py integration."""
    p = subparsers.add_parser("i18n", help="📦 Translation commands")
    i18n_sub = p.add_subparsers(dest="i18n_cmd", metavar="action")

    # Audit command
    audit_p = i18n_sub.add_parser("audit", help="Audit translations for missing keys")
    add_arguments(audit_p)
    audit_p.set_defaults(func=run_from_args)

    # Add command
    add_p = i18n_sub.add_parser("add", help="Add a new translation key to all languages")
    add_p.add_argument("key", help="Translation key (e.g., 'common.save')")
    add_p.add_argument("--en", required=True, help="English translation")
    add_p.add_argument("--it", required=True, help="Italian translation")
    add_p.add_argument("--fr", required=True, help="French translation")
    add_p.add_argument("--es", required=True, help="Spanish translation")
    add_p.set_defaults(func=cmd_add)

    # Remove command
    remove_p = i18n_sub.add_parser("remove", help="Remove a translation key from all languages")
    remove_p.add_argument("key", help="Translation key to remove")
    remove_p.add_argument("-f", "--force", action="store_true", help="Skip confirmation prompt")
    remove_p.set_defaults(func=cmd_remove)

    # Update command
    update_p = i18n_sub.add_parser("update", help="Update a translation in one or more languages")
    update_p.add_argument("key", help="Translation key to update")
    update_p.add_argument("--en", help="English translation")
    update_p.add_argument("--it", help="Italian translation")
    update_p.add_argument("--fr", help="French translation")
    update_p.add_argument("--es", help="Spanish translation")
    update_p.set_defaults(func=cmd_update)

    # Search command
    search_p = i18n_sub.add_parser("search", help="Search for keys/values in translations")
    search_p.add_argument("query", help="Search query (substring match, case-insensitive)")
    search_p.add_argument("-k", "--keys", action="store_true",
                          help="Search only in key names (e.g., 'common.cancel')")
    search_p.add_argument("-v", "--values", action="store_true",
                          help="Search only in translation values")
    search_p.add_argument("-l", "--lang",
                          help="Restrict value search to specific language(s), comma-separated (e.g., 'it' or 'en,es')")
    search_p.set_defaults(func=cmd_search)

    # Tree command
    tree_p = i18n_sub.add_parser("tree", help="Show translation key structure as a tree")
    tree_p.add_argument("prefix", nargs="?", default="",
                        help="Filter tree to keys starting with this prefix (e.g., 'chartSettings')")
    tree_p.add_argument("-d", "--depth", type=int, default=0,
                        help="Max depth to display (0 = unlimited)")
    tree_p.add_argument("--counts", action="store_true",
                        help="Show leaf count per branch")
    tree_p.set_defaults(func=cmd_tree)


if __name__ == "__main__":
    main()
